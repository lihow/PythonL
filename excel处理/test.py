# -*- coding: utf-8 -*-
"""
Created on Wed Dec 27 11:16:17 2017

@author: Lenovo
"""

import xlrd
import xlwt

from openpyxl import Workbook
from openpyxl import load_workbook
import openpyxl


def read_excel():
    # 打开文件
    workbook = xlrd.open_workbook('C:/Users/Lenovo/Desktop/temp/score1.xlsx')
    # 获取所有的sheet
    print(workbook.sheet_names())
    # 根据sheet索引或者名称获取sheet内容
    # sheet2=workbook.sheet_by_index(1)
    sheet2 = workbook.sheet_by_name('Sheet2')

    # sheet的名称，行数，列数
    print('Sheet名称：', sheet2.name, ' 行数：', sheet2.nrows, ' 列数：', sheet2.ncols)

    # 获取整行和整列的值（数组）、
    rows = sheet2.row_values(3)  # 获取第四行内容
    cols = sheet2.col_values(2)  # 获取第三列的内容
    print(rows)
    print(cols)

    # 获取单元格内容
    print(sheet2.cell(1, 0).value, sheet2.cell(1, 0).value.encode('utf-8'))
    # 获取单元格的数据类型
    # ctype : 0 empty,1 string, 2 number, 3 date, 4 boolean, 5 error
    print(sheet2.cell(1, 0).ctype, sheet2.cell(1, 1).ctype)


def merge_sheet():
    # 打开文件
    writebook = xlwt.Workbook('C:/Users/Lenovo/Desktop/temp/score1.xlsx')
    readbook = xlrd.open_workbook('C:/Users/Lenovo/Desktop/temp/score1.xlsx')

    # 得到每个sheet
    sheet1 = readbook.sheet_by_name('Sheet1')
    sheet2 = readbook.sheet_by_name('Sheet2')

    sheet3 = writebook.add_sheet('Sheet3')

    # 得到两个表格的第一列
    sheet1_rows1 = sheet1.row_values(0)
    sheet2_rows2 = sheet2.row_values(0)

    '''
    i=1
    while i< sheet1.nrows:
        value=sheet1_row1[i]
        if sheet2_rows2(value):
            for j in 
    '''


def openpyxl_read():
    # 新建excel
    wb2 = openpyxl.Workbook()
    wb2.save('C:/Users/Lenovo/Desktop/temp/test.xlsx')
    print('新建成功')
    # 读取数据
    wb1 = openpyxl.load_workbook('C:/Users/Lenovo/Desktop/temp/score1.xlsx')
    wb2=openpyxl.load_workbook('C:/Users/Lenovo/Desktop/temp/test.xlsx')
    sheets1 = wb1.get_sheet_names()  # ['Sheet1', 'Sheet2']
    sheets2 = wb2.get_sheet_names()
    # print('读取的excel中的sheet：',sheets1)


    sheet1 = wb1.get_sheet_by_name(sheets1[0])  # 读取sheet1
    sheet2=wb2.get_sheet_by_name(sheets2[0])
    max_row = sheet1.max_row
    max_column = sheet1.max_column
    print('max_row', max_row,'max_column',max_column)


    for m in range(1,max_row+1):
        for n in  range(97,97+max_column):
            n=chr(n)
            i='%s%d'%(n,m)#单元格编号
            cell=sheet1[i].value
            sheet2[i].value=cell
    wb2.save('C:/Users/Lenovo/Desktop/temp/test.xlsx')
    wb1.close()
    wb2.close()




if __name__ == '__main__':
    # read_excel()
    # merge_sheet()
    openpyxl_read()
